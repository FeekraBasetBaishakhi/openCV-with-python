import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

product_per_supplier = {}
total_value_per_supply = {}
product_under_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_numb = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)  #for creating extra

    # Calculation number of products per supplier
    if supplier_name in product_per_supplier:
        current_number_products = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_number_products +1

    else:
        product_per_supplier[supplier_name] =1

    # Calculation total value of the supply
    if supplier_name in total_value_per_supply:
        crnt_total_value = total_value_per_supply.get(supplier_name)
        total_value_per_supply[supplier_name] = crnt_total_value + inventory * price
    else:
        total_value_per_supply[supplier_name] = inventory * price

    # logic products with inventory less than ten

    if inventory < 10:
        product_under_inv[int(product_numb)] = int(inventory)

    # add value for total inventory price to set the .value added
    inventory_price.value = inventory * price


print(" Products per supplier are: \n", product_per_supplier)
print(" Total value per supply is: \n", total_value_per_supply)
print(" Products Under Ten in Inventory Are: \n", product_under_inv)
inventory_file.save("Inventory_with_total_value.xlsx")